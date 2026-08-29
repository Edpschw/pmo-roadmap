#!/usr/bin/env python3
"""Extrai a tabela de produção por campo do pré-sal das planilhas Excel
(.xlsm) que a ANP publica junto do Boletim Mensal da Produção (BMP) e
atualiza data/producao.json — mesmo arquivo de saída de
parse_producao.py (o parser do PDF), mesmo formato de registro por mês.

Por que um script separado do PDF (parse_producao.py): a partir de
~jun/2024 a ANP também disponibiliza o boletim em PDF com uma tabela que
dá pra extrair por texto com segurança, mas os anos anteriores (2017 a
2025) só têm essa garantia via Excel — o PDF mais antigo embaralha a
tabela em blocos por coluna que não dá pra reconstruir com confiança (ver
nota em parse_producao.py, parse_table_columnar). O Excel tem célula de
verdade, sem esse problema: os números vêm exatos, não é texto para
reconhecer por regex.

Uso (um arquivo):
    python3 scripts/parse_producao_xlsm.py caminho/para/boletim.xlsm --url URL

Uso (lote):
    python3 scripts/parse_producao_xlsm.py --dir pasta/ [--url-map arquivo.json]

Sem --url/--url-map, ano e mês são inferidos do PRÓPRIO nome do arquivo
(--dir espera o padrão "{ano}_{nome-original}", ver baixa em lote) — o
nome do arquivo no site da ANP não é consistente entre edições (às vezes
"2020-05-boletim.xlsm", às vezes "tabela-outubro.xlsm", sem ano no nome),
então o ano vem do prefixo "{ano}_" e o mês de um número YYYY-MM- no nome
ou do nome do mês por extenso (em português, sem acento) em algum lugar
do nome do arquivo.
"""
import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from producao_common import (  # noqa: E402
    METRIC_KEYS, REPO_ROOT, DATA_PATH, clean_field_name, load_existing, save, strip_accents, upsert_month,
)

# Nome da aba muda de posição/vizinhas entre edições (nº de abas varia),
# mas o texto "Dados de Produção" (com o número "2." na frente ou não) é
# estável em toda edição observada (2017-2025).
SHEET_NAME_RE = re.compile(r'Dados de Produ', re.IGNORECASE)

# O título exato da tabela varia por edição ("Distribuição da produção
# dos campos do Pré-sal" em edições recentes, "Produção de Campos do
# Pré-sal" em 2017-2018) — casamento solto por conter as duas palavras-
# chave, não pelo texto inteiro.
def is_field_table_title(cell):
    if not cell:
        return False
    s = str(cell).lower()
    return 'campo' in s and 'pré-sal' in s and s.strip().lower().startswith('tabela')

MONTH_NAME_TO_NUM = {
    'janeiro': 1, 'fevereiro': 2, 'marco': 3, 'abril': 4, 'maio': 5, 'junho': 6,
    'julho': 7, 'agosto': 8, 'setembro': 9, 'outubro': 10, 'novembro': 11, 'dezembro': 12,
}


def to_num(v):
    return float(v) if v is not None else 0.0


def find_field_table(wb):
    sheet = next((wb[sn] for sn in wb.sheetnames if SHEET_NAME_RE.search(sn)), None)
    if sheet is None:
        return None
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True), start=1):
        if len(row) > 1 and is_field_table_title(row[1]):
            return sheet, i
    return None


def parse_xlsm(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    found = find_field_table(wb)
    if not found:
        raise RuntimeError('Tabela de produção por campo do pré-sal não encontrada (nenhuma aba/título reconhecido).')
    sheet, title_row = found
    # title_row+1: linha em branco. title_row+2: cabeçalho "Campo |
    # Petróleo | Gás natural | Produção". title_row+3: subcabeçalho
    # "Pré-sal | Pós-sal | ...". Dados começam em title_row+4.
    campos = {}
    for row in sheet.iter_rows(min_row=title_row + 4, max_row=sheet.max_row, values_only=True):
        if len(row) < 8:
            continue
        nome_raw = row[1]
        if nome_raw is None:
            break
        nome_lower = str(nome_raw).strip().lower()
        if nome_lower.startswith('total'):
            break
        nome = clean_field_name(nome_raw)
        if not nome:
            continue
        vals = [to_num(row[k]) for k in range(2, 8)]
        entry = dict(zip(METRIC_KEYS, vals))
        if nome in campos:
            for k in METRIC_KEYS:
                campos[nome][k] += entry[k]
        else:
            campos[nome] = entry
    if not campos:
        raise RuntimeError('Tabela encontrada, mas sem linhas de dado reconhecidas.')
    return campos


def infer_ano_mes(filename, url=None):
    source = url or filename
    ano_match = re.search(r'/((?:19|20)\d{2})/', source) or re.search(r'^((?:19|20)\d{2})_', filename)
    if not ano_match:
        raise RuntimeError(f'Não consegui inferir o ano a partir de "{filename}".')
    ano = int(ano_match.group(1))

    mes_match = re.search(r'(?:19|20)\d{2}[_-](\d{2})[_-]', filename)
    if mes_match:
        return ano, int(mes_match.group(1))

    normalized = strip_accents(filename.lower())
    for name, num in MONTH_NAME_TO_NUM.items():
        if name in normalized:
            return ano, num
    raise RuntimeError(f'Não consegui inferir o mês a partir de "{filename}".')


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('xlsm', nargs='?', help='Caminho de um único arquivo .xlsm')
    parser.add_argument('--url', help='URL original (modo arquivo único)')
    parser.add_argument('--dir', help='Pasta com vários .xlsm (modo lote)')
    parser.add_argument('--url-map', help='JSON {nome-do-arquivo: URL} pro modo lote')
    args = parser.parse_args()

    existing = load_existing()

    if args.dir:
        url_map = {}
        if args.url_map:
            url_map = json.loads(Path(args.url_map).read_text(encoding='utf-8'))
        files = sorted(Path(args.dir).glob('*.xlsm'))
        ok, failed = [], []
        for path in files:
            url = url_map.get(path.name)
            try:
                ano, mes = infer_ano_mes(path.name, url)
                campos = parse_xlsm(path)
                upsert_month(existing, ano, mes, campos, url)
                ok.append((path.name, ano, mes, len(campos)))
            except Exception as e:
                failed.append((path.name, str(e)))
        save(existing)
        print(f'OK: {len(ok)}/{len(files)} planilhas importadas para {DATA_PATH.relative_to(REPO_ROOT)}')
        for name, ano, mes, n in ok:
            print(f'  {mes:02d}/{ano}  {name}  ({n} campos)')
        for name, reason in failed:
            print(f'  FALHOU {name}: {reason}')
        return

    if not args.xlsm:
        parser.print_help()
        sys.exit(1)
    ano, mes = infer_ano_mes(Path(args.xlsm).name, args.url)
    campos = parse_xlsm(args.xlsm)
    upsert_month(existing, ano, mes, campos, args.url)
    save(existing)
    print(f'OK: {mes:02d}/{ano} — {len(campos)} campos gravados em {DATA_PATH.relative_to(REPO_ROOT)}')


if __name__ == '__main__':
    main()
